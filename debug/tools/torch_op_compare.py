import argparse
import ast
import json
import os.path
import time
from queue import Queue
import numpy as np

from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.workbook import Workbook

GPU = 0
NPU = 1
NA = 'N/A'
LIMIT_KERNEL = 3
GPU_OP_NAME = 'GPU Operator Name'
GPU_KERNEL_NAME = 'GPU Kernel Name'
GPU_DUR = 'GPU Duration(us)'
NPU_OP_NAME = 'NPU Operator Name'
NPU_KERNEL_NAME = 'NPU Kernel Name'
NPU_TASK_ID = 'NPU Task Id'
NPU_KERNEL_TYPE = 'NPU Kernel Type'
NPU_DUR = 'NPU Duration(us)'
DIFF = 'DIFF: (sum(NPU Duration)-sum(GPU Duration))/sum(GPU Duration)'
OP_NAME_FILTER = 'Operator Name Filter'
DIFF_FILTER = 'DIFF Filter'
HEADERS = [
    GPU_OP_NAME, GPU_KERNEL_NAME, GPU_DUR, NPU_OP_NAME, NPU_KERNEL_NAME, NPU_TASK_ID, NPU_KERNEL_TYPE, NPU_DUR, DIFF,
    OP_NAME_FILTER, DIFF_FILTER
]
FILL_DICT = {
    GPU_OP_NAME: PatternFill("solid", fgColor='003366FF'), GPU_KERNEL_NAME: PatternFill("solid", fgColor='003366FF'),
    GPU_DUR: PatternFill("solid", fgColor='003366FF'), NPU_OP_NAME: PatternFill("solid", fgColor='0033CCCC'),
    NPU_KERNEL_NAME: PatternFill("solid", fgColor='0033CCCC'), NPU_TASK_ID: PatternFill("solid", fgColor='0033CCCC'),
    NPU_KERNEL_TYPE: PatternFill("solid", fgColor='0033CCCC'), NPU_DUR: PatternFill("solid", fgColor='0033CCCC'),
    DIFF: PatternFill("solid", fgColor='00FF0000'), OP_NAME_FILTER: PatternFill("solid", fgColor='00FFFF00'),
    DIFF_FILTER: PatternFill("solid", fgColor='00FFFF00')
}
COLUMN_WIDTH = {GPU_OP_NAME: 50, GPU_KERNEL_NAME: 25, GPU_DUR: 25, NPU_OP_NAME: 50, NPU_KERNEL_NAME: 25,
                NPU_TASK_ID: 20, NPU_KERNEL_TYPE: 25, NPU_DUR: 25, DIFF: 25, OP_NAME_FILTER: 25, DIFF_FILTER: 25}
BORDER = Border(top=Side(border_style="thin", color='00000000'),
                left=Side(border_style="thin", color='00000000'),
                right=Side(border_style="thin", color='00000000'),
                bottom=Side(border_style="thin", color='00000000'))


class TorchOpNode:
    def __init__(self, event=None, parent_node=None):
        self._event = event
        self._parent_node = parent_node
        self._child_nodes = []
        self._kernel_list = []
        self._kernel_num = 0

    @property
    def start_time(self):
        return self._event.get("ts", 0)

    @property
    def end_time(self):
        return self._event.get("ts", 0) + self._event.get("dur", 0)

    @property
    def name(self):
        return self._event.get("name", NA)

    @property
    def parent(self):
        return self._parent_node

    @property
    def child_nodes(self):
        return self._child_nodes

    @property
    def kernel_list(self):
        return self._kernel_list

    @property
    def kernel_num(self):
        return self._kernel_num

    def add_child_node(self, child_node):
        self._child_nodes.append(child_node)

    def set_kernel_list(self, kernel_list: list):
        self._kernel_list = kernel_list

    def add_kernel_num(self, kernel_num: int):
        self._kernel_num += kernel_num

    def is_step_profiler(self) -> bool:
        return self.name.find("ProfilerStep#") != -1

    def is_skipped_op(self, skipped_op_list: list) -> bool:
        return self.name in skipped_op_list


class TreeBuilder:
    @classmethod
    def build_tree(cls, event_list: list, flow_kernel_dict: dict) -> TorchOpNode:
        root_node = TorchOpNode()
        event_list.sort(key=lambda x: x.get("ts", 0))
        last_node = root_node
        for event in event_list:
            kernel_list = flow_kernel_dict.get(event.get("ts", 0), [])
            node_queue = Queue()
            node_queue.put(last_node)
            while not node_queue.empty():
                compare_node = node_queue.get()
                if compare_node == root_node or event.get("ts", 0) < compare_node.end_time:
                    tree_node = TorchOpNode(event, compare_node)
                    compare_node.add_child_node(tree_node)
                    if kernel_list:
                        tree_node.set_kernel_list(kernel_list)
                    last_node = tree_node
                    break
                node_queue.put(compare_node.parent)
        return root_node

    @classmethod
    def mark_kernel_num(cls, root_node: TorchOpNode, flow_kernel_dict: dict):
        for ts, kernel_list in flow_kernel_dict.items():
            curr_node = root_node
            while curr_node.child_nodes:
                for node in curr_node.child_nodes:
                    if node.start_time <= ts <= node.end_time:
                        node.add_kernel_num(len(kernel_list))
                        curr_node = node
                        break

    @classmethod
    def get_total_kernels(cls, root_node: TorchOpNode) -> list:
        result_list = []
        node_queue = Queue()
        for child_node in root_node.child_nodes:
            node_queue.put(child_node)
        while not node_queue.empty():
            tree_node = node_queue.get()
            result_list.extend(tree_node.kernel_list)
            for child_node in tree_node.child_nodes:
                node_queue.put(child_node)
        return result_list


def read_json_file(file_path: str, type_id: int) -> any:
    event_list = []
    flow_kernel_dict = {}
    if not os.path.isfile(file_path):
        raise RuntimeError(f"File not exists: {file_path}")
    try:
        with open(file_path, "rt") as file:
            json_data = json.loads(file.read())
    except Exception:
        raise RuntimeError(f"Can't read file: {file_path}")
    flow_start_dict, flow_end_dict, event_dict = {}, {}, {}
    flow_cat = ("async_gpu", "ac2g", "async_npu")
    total_events = json_data.get("traceEvents", []) if type_id == GPU else json_data
    for event in total_events:
        if event.get("cat") == "cpu_op" or event.get("cat") in ("Runtime", "cuda_runtime"):
            event_list.append(event)
        elif event.get("cat") in flow_cat and event.get("ph") == "s":
            flow_start_dict[event.get("id")] = event
        elif event.get("cat") in flow_cat and event.get("ph") == "f":
            flow_end_dict[event.get("id")] = event
        elif type_id == GPU and event.get("cat", "").capitalize() == "Kernel".capitalize():
            event_dict["{}-{}-{}".format(event.get("pid"), event.get("tid"), event.get("ts"))] = event
        elif type_id == NPU and event.get("ph") != "f":
            event_dict["{}-{}-{}".format(event.get("pid"), event.get("tid"), event.get("ts"))] = event

    for flow_id, start_flow in flow_start_dict.items():
        end_flow = flow_end_dict.get(flow_id)
        if end_flow is None:
            continue
        kernel_event = event_dict.get("{}-{}-{}".format(end_flow.get("pid"), end_flow.get("tid"), end_flow.get("ts")))
        if kernel_event is None:
            continue
        flow_kernel_dict.setdefault(start_flow.get("ts"), []).append(kernel_event)
    return event_list, flow_kernel_dict


def get_top_layer_apis(file_path: str, type_id: int, max_kernel_num: int) -> any:
    event_list, flow_kernel_dict = read_json_file(file_path, type_id)
    root_node = TreeBuilder.build_tree(event_list, flow_kernel_dict)
    if max_kernel_num is not None:
        TreeBuilder.mark_kernel_num(root_node, flow_kernel_dict)
    level1_child_nodes = root_node.child_nodes
    if not level1_child_nodes:
        raise RuntimeError(f"Can't find any torch op in the file: {file_path}")
    result_data = []
    for level1_node in level1_child_nodes:
        if level1_node.is_step_profiler():
            result_data.extend(level1_node.child_nodes)
        else:
            result_data.append(level1_node)
    return result_data


def compare(gpu_top_layer_apis: list, npu_top_layer_apis: list, op_name_map: dict) -> list:
    result_data = []
    npu_len, gpu_len = len(npu_top_layer_apis), len(gpu_top_layer_apis)
    dp = [[0] * (gpu_len + 1) for _ in range(npu_len + 1)]
    for npu_index in range(1, npu_len + 1):
        for gpu_index in range(1, gpu_len + 1):
            if op_name_map.get(npu_top_layer_apis[npu_index - 1].name,
                               npu_top_layer_apis[npu_index - 1].name) == op_name_map.get(
                gpu_top_layer_apis[gpu_index - 1].name, gpu_top_layer_apis[gpu_index - 1].name):
                dp[npu_index][gpu_index] = dp[npu_index - 1][gpu_index - 1] + 1
            else:
                dp[npu_index][gpu_index] = max(dp[npu_index][gpu_index - 1], dp[npu_index - 1][gpu_index])
    matched_op = []
    npu_index, gpu_index = npu_len, gpu_len
    while npu_index > 0 and gpu_index > 0:
        if op_name_map.get(npu_top_layer_apis[npu_index - 1].name,
                           npu_top_layer_apis[npu_index - 1].name) == op_name_map.get(
            gpu_top_layer_apis[gpu_index - 1].name, gpu_top_layer_apis[gpu_index - 1].name):
            matched_op.append([npu_index - 1, gpu_index - 1])
            npu_index -= 1
            gpu_index -= 1
            continue
        if dp[npu_index][gpu_index - 1] > dp[npu_index - 1][gpu_index]:
            gpu_index -= 1
        else:
            npu_index -= 1
    if not matched_op:
        matched_gpu_index_list = []
    else:
        matched_op.reverse()
        matched_op = np.array(matched_op)
        matched_gpu_index_list = list(matched_op[:, 1])
    curr_npu_index = 0
    for gpu_index, gpu_api_node in enumerate(gpu_top_layer_apis):
        if gpu_index not in matched_gpu_index_list:
            result_data.append([gpu_api_node, None])
            continue
        matched_npu_index = matched_op[matched_gpu_index_list.index(gpu_index), 0]
        for npu_index in range(curr_npu_index, matched_npu_index):
            result_data.append([None, npu_top_layer_apis[npu_index]])
        result_data.append([gpu_api_node, npu_top_layer_apis[matched_npu_index]])
        curr_npu_index = matched_npu_index + 1
    if curr_npu_index < len(npu_top_layer_apis):
        for npu_index in range(curr_npu_index, len(npu_top_layer_apis)):
            result_data.append([None, npu_top_layer_apis[npu_index]])
    return result_data


def create_data(gpu_api_node: TorchOpNode, npu_api_node: TorchOpNode) -> list:
    result_data = []
    gpu_kernel_list = TreeBuilder.get_total_kernels(gpu_api_node) if gpu_api_node is not None else []
    npu_kernel_list = TreeBuilder.get_total_kernels(npu_api_node) if npu_api_node is not None else []
    if not gpu_kernel_list or not npu_kernel_list:
        diff = NA
    else:
        gpu_total_dur = sum([kernel.get("dur", 0) for kernel in gpu_kernel_list])
        npu_total_dur = sum([kernel.get("dur", 0) for kernel in npu_kernel_list])
        diff = (npu_total_dur - gpu_total_dur) / gpu_total_dur
    gpu_op_name = NA if gpu_api_node is None else gpu_api_node.name
    npu_op_name = NA if npu_api_node is None else npu_api_node.name
    op_name = gpu_op_name if npu_op_name == NA else npu_op_name
    gpu_kernel_num, npu_kernel_num = len(gpu_kernel_list), len(npu_kernel_list)
    if gpu_kernel_num == 0 and npu_kernel_num == 0:
        data = [gpu_op_name, NA, NA, npu_op_name, NA, NA, NA, NA, NA, op_name]
        result_data.append(data)
        return result_data
    for index in range(max(gpu_kernel_num, npu_kernel_num)):
        gpu_kernel_name, gpu_kernel_dur, npu_kernel_name, npu_task_id, npu_kernel_type, npu_kernel_dur = NA, NA, NA, NA, NA, NA
        if index < gpu_kernel_num:
            gpu_kernel = gpu_kernel_list[index]
            gpu_kernel_name = gpu_kernel.get("name")
            gpu_kernel_dur = gpu_kernel.get("dur")
        if index < npu_kernel_num:
            npu_kernel = npu_kernel_list[index]
            npu_kernel_name = npu_kernel.get("name")
            npu_kernel_dur = npu_kernel.get("dur")
            npu_kernel_type = npu_kernel.get("args", {}).get("Task Type")
            npu_task_id = npu_kernel.get("args", {}).get("Task Id")
        data = [gpu_op_name, gpu_kernel_name, gpu_kernel_dur, npu_op_name, npu_kernel_name, npu_task_id,
                npu_kernel_type, npu_kernel_dur, diff, op_name]
        result_data.append(data)
    return result_data


def drill_down(compare_result_data: list, max_kernel_num: int, op_name_map: dict) -> list:
    result_data = []
    for data in compare_result_data:
        gpu_api = data[0] if data[0] else TorchOpNode()
        npu_api = data[1] if data[1] else TorchOpNode()
        if max(gpu_api.kernel_num, npu_api.kernel_num) <= max_kernel_num:
            result_data.append(data)
            continue
        result_data.extend(compare(gpu_api.child_nodes, npu_api.child_nodes, op_name_map))
    return result_data


def have_to_drill_down(compare_result_data: list, max_kernel_num: int) -> bool:
    for data in compare_result_data:
        gpu_api = data[0] if data[0] else TorchOpNode()
        npu_api = data[1] if data[1] else TorchOpNode()
        if max(gpu_api.kernel_num, npu_api.kernel_num) > max_kernel_num:
            return True
    return False


def main():
    parser = argparse.ArgumentParser(description="Compare trace of GPU and NPU")
    parser.add_argument("gpu_trace_path", help="GPU trace file path")
    parser.add_argument("npu_trace_path", help="NPU trace file path")
    parser.add_argument("--output_path", help="性能数据比对结果的存放路径")
    parser.add_argument("--max_kernel_num", type=int, help="每个torch op的kernel数量限制")
    parser.add_argument("--op_name_map", type=ast.literal_eval, default={},
                        help="配置GPU OP与NPU OP等价的名称映射关系，以字典的形式传入")
    args = parser.parse_args()
    if args.max_kernel_num is not None and args.max_kernel_num <= LIMIT_KERNEL:
        raise RuntimeError(f"Invalid param, --max_kernel_num has to be greater than {LIMIT_KERNEL}")
    if not isinstance(args.op_name_map, dict):
        raise RuntimeError("Invalid param, --op_name_map must be dict, for example: --op_name_map={'name1':'name2'}")
    gpu_top_layer_apis = get_top_layer_apis(args.gpu_trace_path, GPU, args.max_kernel_num)
    npu_top_layer_apis = get_top_layer_apis(args.npu_trace_path, NPU, args.max_kernel_num)
    compare_result_data = compare(gpu_top_layer_apis, npu_top_layer_apis, args.op_name_map)

    if args.max_kernel_num is not None:
        while have_to_drill_down(compare_result_data, args.max_kernel_num):
            compare_result_data = drill_down(compare_result_data, args.max_kernel_num, args.op_name_map)

    dir_path = args.output_path if args.output_path else "./"
    file_name = "torch_op_compare_{}.xlsx".format(time.strftime("%Y%m%d%H%M%S", time.localtime(time.time())))
    result_file_path = os.path.join(dir_path, file_name)

    wb = Workbook()
    ws = wb.create_sheet("CompareResult", 0)
    ws.sheet_properties.tabColor = '00CED1'
    # write headers
    for col_index in range(len(HEADERS)):
        header_name = HEADERS[col_index]
        ws.cell(row=1, column=col_index + 1).value = header_name
        dim = ws.cell(row=1, column=col_index + 1).coordinate
        ws.column_dimensions[dim[0]].width = COLUMN_WIDTH.get(header_name)
        ws.cell(row=1, column=col_index + 1).font = Font(name='Arial', bold=True)
        ws.cell(row=1, column=col_index + 1).fill = FILL_DICT.get(header_name)
        ws.cell(row=1, column=col_index + 1).border = BORDER
    # write lines
    start_row_index = 2
    for data in compare_result_data:
        rows = create_data(data[0], data[1])
        row_number = 0
        for row in rows:
            row_index = start_row_index + row_number
            ws.cell(row=row_index, column=len(row) + 1).border = BORDER
            for index, value in enumerate(row):
                if index == HEADERS.index(DIFF):
                    ws.cell(row=row_index, column=index + 1).number_format = '0.00%'
                    if value != NA and value < 0:
                        ws.cell(row=row_index, column=index + 1).fill = PatternFill("solid", fgColor='0000FF00')
                        ws.cell(row=row_index, column=index + 3).fill = PatternFill("solid", fgColor='0000FF00')
                    if value != NA and value >= 0:
                        ws.cell(row=row_index, column=index + 1).fill = PatternFill("solid", fgColor='00FF0000')
                        ws.cell(row=row_index, column=index + 3).fill = PatternFill("solid", fgColor='00FF0000')
                if index in (HEADERS.index(GPU_OP_NAME), HEADERS.index(NPU_OP_NAME)):
                    ws.cell(row=row_index, column=index + 1).font = Font(name='Arial', bold=True)
                else:
                    ws.cell(row=row_index, column=index + 1).font = Font(name='Arial')
                ws.cell(row=row_index, column=index + 1).value = value
                ws.cell(row=row_index, column=index + 1).border = BORDER
            row_number += 1
        if row_number > 1:
            # 合并单元格
            for col_index in [HEADERS.index(GPU_OP_NAME), HEADERS.index(NPU_OP_NAME), HEADERS.index(DIFF)]:
                ws.merge_cells(start_row=start_row_index, start_column=col_index + 1,
                               end_row=start_row_index + row_number - 1, end_column=col_index + 1)
        start_row_index = start_row_index + row_number

    wb.save(result_file_path)
    wb.close()


if __name__ == "__main__":
    main()
